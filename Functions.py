import random
import string
import os
import openpyxl

class TextColors:
    RESET = "\033[0m"
    RED = "\033[91m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    BLUE = "\033[94m"
    MAGENTA = "\033[95m"
    CYAN = "\033[96m"
    
# List of Hindi words transliterated to English (you can extend this list)
hindi_words_transliterated = [
    # ... (same as in your original code)
    "Sundar", "Sukha", "Pyaara", "Chhota", "Mota", "Bura", "Khoobsurat",
    "Swasth", "Garib", "Ameer", "Nanga", "Jhulta", "Majedar", "Garam",
    "Thanda", "Naya", "Purana", "Lal", "Hara", "Neela", "Safed", "Kaala",
    "Sukha", "Dhaniya", "Lamba", "Samridh", "Sachcha", "Jhoota", "Sharif",
    "Tez", "Teekha", "Udta", "Langda", "Jhagdalu", "Budbak", "Gaddar", "Bhutiya",
    "Chulbul","Bhukkad","Peela","Lambu","Namkeen","Shaana"
]

hindi_words_transliterated2 = [
    # ... (same as in your original code)
     "Kanya", "Talwar", "Teer", "Tyagi", "Chhatri", "Pappu","Hira",
    "Maal", "Rasta", "Mazdur", "kitab", "kaagaz", "Tubelight","Moza",
    "kursi", "botal", "raja", "don", "bhai", "Sipahi","Bijli","Jhola",
    "Shakeel", "gunda", "Aurat", "Panchi", "Chaman","Mogambo","Pustak"
    "Khidki", "Darwaza", "Mahal", "Gulaab","Kachra","Paanwala","Ghoda"
]

def replace_characters(word):
    # ... (same as in your original code)
    # Replace 's' with '$', 'o' with '0'
    word = word.replace('s', '$')
    word = word.replace('o', '0')
    return word

def generate_strong_password(password_length):
    # ... (same as in your original code)
    if password_length < 6:
        return "Password length is too short (minimum length is 6 characters)."

    while True:
        word1 = random.choice(hindi_words_transliterated)
        word2 = random.choice(hindi_words_transliterated2)
        if word1 != word2:
            break

    word1 = replace_characters(word1).capitalize()
    word2 = replace_characters(word2).capitalize()
    readable = f"{word1} {word2}"
    
    digit = random.choice(string.digits)
    special_char = random.choice(string.punctuation)

    password = [word1, word2, digit, special_char]

    indicestoshuffle = [0, 2, 3]
    shuffled_elements = random.sample([password[i] for i in indicestoshuffle], len(indicestoshuffle))
    random.shuffle(shuffled_elements)
    for i, index in enumerate(indicestoshuffle):
        password[index] = shuffled_elements[i]

    password = ''.join(map(str, password))
    readable = f"{word1} {word2}"

    return password, readable

def generate_passwords(num_passwords=5):
    # ... (same as in your original code)
    password_length = 11
    passwords = []
    readables = []

    for _ in range(num_passwords):
        password, readable = generate_strong_password(password_length)
        passwords.append(password)
        readables.append(readable)
    
    print_Password(passwords,readables)
    return passwords

def print_Password(passw,readables):
    print("")
    print(TextColors.YELLOW +"======= Generating Memorable Strong Passwords ======="+TextColors.RESET)
    print("")
    for i, each in enumerate(passw):
        print(f"Strong Password: {i + 1}. {readables[i]} - {each}")



def CreateExcel(rowdata):
    # ... (same as in your original code)
    # Check if the Excel file exists
    os.chdir(r"G:\PY\passowordgen")
    file_path = 'PasswordRecords.xlsx' # Use the same file_path variable
    if os.path.exists(file_path):
        # File exists, open it
        workbook = openpyxl.load_workbook(file_path)
        # print(f"Opening existing Excel file: {file_path}")
    else:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    # Create a new worksheet (if needed)
    worksheet = workbook.active

    # Check if the first row is already filled
    first_row_empty = all(cell.value is None for cell in worksheet[1])

    if first_row_empty:
        # Add your header data to the first row
        header = ["Password","Application_Account"]
        for i, value in enumerate(header, start=1):
            worksheet.cell(row=1, column=i, value=value)
        worksheet.append(rowdata)
    else:
        # Add your data to the first row
        worksheet.append(rowdata)

    workbook.save("PasswordRecords.xlsx")
    workbook.close()
    return

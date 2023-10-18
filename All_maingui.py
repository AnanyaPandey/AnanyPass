# -*- coding: utf-8 -*-
"""
Created on Mon Oct 16 16:40:15 2023

@author: pandey
GUI - Application to Generate a Remorable Hindi Word Strong Password
Application Generates the list of strong passwords in a simple click.
IF user doesnt like any of those, or want to see another list it allows to re generate list with new passwords
User then has an option to select and Copy it and save it in Excel File for Records
It is recommended not to use a strong memorable password to open the Excel passwird

"""
import pickle
import random
import string
from pathlib import Path
import PySimpleGUI as sg
import pyperclip
import openpyxl
import re
from sys import exit

ROOTDIR = Path('G:/py/Ananypass')

# List of Hindi words transliterated to English (you can extend this list)
hindi_words_transliterated = [
    # ... (same as in your original code)
    "Sundar", "Sukha", "Pyaara", "Chhota", "Mota", "Bura", "Khoobsurat",
    "Swasth", "Garib", "Ameer", "Nanga", "Jhulta", "Majedar", "Garam",
    "Thanda", "Naya", "Purana", "Lal", "Hara", "Neela", "Safed", "Kala",
    "Sukha", "Dhaniya", "Lamba", "Samridh", "Sachcha", "Jhoota", "Sharif",
    "Tez", "Teekha", "Udta", "Langda", "Jhagdalu", "Budbak", "Gaddar", "Bhutiya",
    "Chulbul","Bhukkad","Peela","Lambu","Namkeen","Shaana"
]

hindi_words_transliterated2 = [
    # ... (same as in your original code)
     "Kanya", "Talwar", "Teer", "Tyagi", "Chhatri", "Pappu","Hira",
    "Maal", "Rasta", "Mazdur", "kitab", "kaagaz", "Tubelight","Moza",
    "kursi", "botal", "raja", "don", "bhai", "Sipahi","Bijli","Jhola",
    "Shakeel", "gunda", "Aurat", "Panchi", "Chaman","Mogambo","Pustak"
    "Khidki", "Darwaza", "Mahal", "Gulaab","Kachra","Paanwala","Ghoda"
]

def replace_characters(word):
    # ... (same as in your original code)
    # Replace 's' with '$', 'o' with '0'
    word = word.replace('s', '$')
    word = word.replace('o', '0')
    return word

def generate_strong_password(password_length):
    # ... (same as in your original code)
    if password_length < 7:
        return "Password length is too short (minimum length is 7 characters)."

    while True:
        word1 = random.choice(hindi_words_transliterated)
        word2 = random.choice(hindi_words_transliterated2)
        if word1 != word2:
            break

    readable = f"{word1} {word2}"
    word1 = replace_characters(word1).capitalize()
    word2 = replace_characters(word2).capitalize()


    digit = random.choice(string.digits)
    special_char = random.choice(string.punctuation)

    password = [word1, word2, digit, special_char]

    indicestoshuffle = [0, 2, 3]
    shuffled_elements = random.sample([password[i] for i in indicestoshuffle], len(indicestoshuffle))
    random.shuffle(shuffled_elements)
    for i, index in enumerate(indicestoshuffle):
        password[index] = shuffled_elements[i]

    password = ''.join(map(str, password))
    # readable = f"{word1} {word2}"

    return password, readable

def generate_passwords(num_passwords=5):
    # ... (same as in your original code)
    password_length = 11
    passwords = []
    readables = []

    for _ in range(num_passwords):
        password, readable = generate_strong_password(password_length)
        passwords.append(password)
        readables.append(readable)

    # print_Password(passwords,readables)
    return passwords,readables

def pull_hindi_words(filename):
    # filepaths = ROOTDIR.glob("Some*.txt")
    # for each in filepaths:
    fname=Path(filename)
    with open(fname,'r') as file :
        passw = file.readlines()
    return passw

def generate_user_pass(user_pass_words,password_length=10):
    # ... (same as in your original code)
    if password_length < 7:
        return "Password length is too short (minimum length is 7 characters)."

    while True:
        word1 = random.choice(user_pass_words)
        word2 = random.choice(user_pass_words)
        if word1 != word2:
            break

    readable = f"{word1} {word2}"
    word1 = replace_characters(word1).capitalize()
    word2 = replace_characters(word2).capitalize()


    digit = random.choice(string.digits)
    special_char = random.choice(string.punctuation)
    password = [word1, word2, digit, special_char]

    indicestoshuffle = [0, 2, 3]
    shuffled_elements = random.sample([password[i] for i in indicestoshuffle], len(indicestoshuffle))
    random.shuffle(shuffled_elements)
    for i, index in enumerate(indicestoshuffle):
        password[index] = shuffled_elements[i]

    password = ''.join(map(str, password))
    # readable = f"{word1} {word2}"

    return password, readable
def generate_userpass(user_pass_words,num_passwords=5):
    # ... (same as in your original code)
    password_length = 11
    passwords = []
    readables = []

    for _ in range(num_passwords):
        password, readable = generate_user_pass(user_pass_words)
        passwords.append(password)
        readables.append(readable)

    # print_Password(passwords,readables)
    return passwords,readables

def CreateExcel_try(rowdata, location):
    # Create a Path object for the target directory
    target_dir = Path(location)
    filename = 'PasswordRecords.xlsx'
    file_path = target_dir / filename
    if file_path.exists():
        # File exists, open it
        workbook = openpyxl.load_workbook(file_path)
    else:
        # Create a new Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    # Create a new worksheet (if needed)
    worksheet = workbook.active

    # Check if the first row is already filled
    first_row_empty = all(cell.value is None for cell in worksheet[1])

    if first_row_empty:
        # Add your header data to the first row
        header = ["Password", "Application_Account"]
        for i, value in enumerate(header, start=1):
            worksheet.cell(row=1, column=i, value=value)

    # Add your data to the worksheet
    worksheet.append(rowdata)

    # Save and close the workbook with the file_path
    workbook.save(file_path)
    workbook.close()


def check_password_strength(password):
    # Check if the password length is at least 8 characters
    if len(password) < 8:
        return "Weak: Password is too short (minimum length is 8 characters)."

    # Check if the password contains at least one uppercase letter
    if not any(char.isupper() for char in password):
        return "Weak: Password must contain at least one uppercase letter."

    # Check if the password contains at least one lowercase letter
    if not any(char.islower() for char in password):
        return "Weak: Password must contain at least one lowercase letter."

    # Check if the password contains at least one digit
    if not any(char.isdigit() for char in password):
        return "Weak: Password must contain at least one digit."

    # Check if the password contains at least one special character
    if not any(char in string.punctuation for char in password):
        return "Weak: Password must contain at least one special character."

    # Check if the password is not a common password or a dictionary word
    if is_common_password(password):
        return "Weak: Password is too common."

    if re.search(r'(.)\1{2,}', password):
        return "Weak Password : It has Reccuring Characters"

    TimetoCrack = estimate_crack_time(password)
    Formatted_TTK = format_number(TimetoCrack)

    # Password meets all complexity criteria
    passstrength = "Strong"

    return [passstrength, Formatted_TTK]


def format_number(number):
    if number < 1000:
        return str(number)
    elif number < 1000000:
        return f"{number / 1000:.1f} Thousand"
    elif number < 1000000000:
        return f"{number / 10000000:.1f} Million"
    elif number < 1000000000000:
        return f"{number / 10000000000:.1f} Billion"
    else:
        return f"{number / 1000000000000:.1f} Trillion"


def is_common_password(password):
    try:
        # Read common passwords from a local file
        filepath = "G:\\PY\\passowordgen\\Ananypass\\common_passwords.txt"
        with open(filepath, 'r') as file:
            common_passwords = set(file.read().splitlines())

        # Check if the password is in the list of common passwords
        return password in common_passwords
    except FileNotFoundError:
        print("Common passwords file not found.")
        return False


def estimate_crack_time(password, attempts_per_second=100000000000):
    # Calculate the number of possible combinations
    character_set_size = 94  # Printable ASCII characters
    password_length = len(password)
    combinations = character_set_size ** password_length

    # Calculate time to crack in seconds
    seconds_to_crack = combinations / attempts_per_second

    # Convert seconds to years
    seconds_in_a_year = 60 * 60 * 24 * 365.25
    years_to_crack = seconds_to_crack / seconds_in_a_year
    # years_to_crack = years_to_crack/100

    return round(years_to_crack)
def main():
    LabelSelfPass = sg.Text("You want Password generator\n"
                            "to consider custom words\n "
                            "provide it here each line")
    label1 = sg.Text("Enter No.of Passwords:")
    input_box_no_of_passwords = sg.InputText(tooltip="# Password",
                                             key="NofPassword",
                                             font=('Calibri', 8),
                                             size=(20, 1))
    gen_button = sg.Button("Generate", key="generate")
    listbox_pass = sg.Listbox(values="",
                              key="listofpasswords",
                              enable_events=True,
                              size=[20, 13])
    listbox_reads = sg.Listbox(values="",
                               key="listofreads",
                               size=[20, 13])
    inputbox_acount = sg.InputText(tooltip="Save for Account",
                                   key="Account",
                                   font=('Calibri', 8),
                                   size=(20,1))
    save_button = sg.Button("Save Record", key="save")
    copy_button = sg.Button("Copy", key="copy")
    label_account = sg.Text("Save it for Account:")
    label2 = sg.Text("Generate Password and Copy it to use", key="Strength")
    label3 = sg.Text("You can Enter an Account and Store it in Excel", key="Updates")
    label4 = sg.Text("", key="Message")
    sep = sg.VSeparator()
    seph = sg.HSeparator()
    inputbox_multi = sg.Multiline("", enable_events=True, key='multi',
                                  size=(20,15),
                                  justification='left',
                                  background_color="lightblue")
    use_button = sg.Button("Use These", key="user_pass",
                           size=(10,1))
    inputbox_folder = sg.Input(tooltip="Chosen Folder to save the Passwords",
                               key="folder_to_save",
                               size=(33,1))
    button_folder = sg.FolderBrowse("Select Folder",key="Folder")
    Lbl = sg.Text("",key="Success")
    # filler1 = sg.Text("")
    # filler2 = sg.Text("")
    left_column_content = [[LabelSelfPass],
                           [use_button],
                           [inputbox_multi],]

    # Prepare the widgets for the right column
    right_column_content = [[label1,input_box_no_of_passwords,gen_button],
                            [listbox_reads,listbox_pass,copy_button],
                            [label2],
                            [label3],
                            [seph],
                            [label4],
                            [inputbox_folder,button_folder],
                            [label_account,inputbox_acount,save_button],
                            [Lbl]]

    left_column = sg.Column(left_column_content)
    right_column = sg.Column(right_column_content)

    LAYOUT = [[left_column, sep, right_column]]

    window = sg.Window(title="Memorable Password Generator",
                       layout=LAYOUT,
                       font=('calibri', 10),
                       size=(670, 430))

    while True:
        event, value = window.read()
        if event == sg.WIN_CLOSED:
            window.close()
            exit()

        print(event)
        print(value)

        if event == "generate":
            try:
                Passwords_to_generate = int(value['NofPassword'])
                if Passwords_to_generate >99 :
                    sg.popup("Enter a Smaller Number, You can Regenerate if you need.")
                    window['NofPassword'].update("")
                    continue
                passwords, readables = generate_passwords(Passwords_to_generate)
                window['listofpasswords'].update(values=passwords)
                window['listofreads'].update(values=readables)
                window['folder_to_save'].update(value="")
            except ValueError:
                sg.popup("Enter Correct Number of passwords to Generate", font=('calibri', 12))
        elif event == "copy":
            try:
                to_copy = value['listofpasswords'][0]
                pyperclip.copy(to_copy)
                copied_text = pyperclip.paste()
                if copied_text == to_copy:
                    window['Updates'].update(value="Succesfully Copied to the Clipboard",text_color="navy")
                PasswordStrength = check_password_strength(copied_text)
                Message = f"Your Selected Password is {PasswordStrength[0]}, it will take {PasswordStrength[1]} years to Crack"
                window['Strength'].update(value=Message)
                window['Message'].update(value="Consider Saving it in Records - You only need to select Folder Once")
                window['folder_to_save'].update(value="")
            except IndexError:
                sg.popup("Select a Password First", font=('calibri', 12))

        elif event == "user_pass":
            try:
                list_to_consider = value['multi'].split('\n')
                print(type(list_to_consider))
                print(list_to_consider)
                passwords,readables = generate_userpass(list_to_consider)
                window['listofpasswords'].update(values=passwords)
                window['listofreads'].update(values=readables)
            except IndexError:
                sg.popup("Provide Custom Words, or use Common words in the next section")

        elif event == "Folder":
            pass
        elif event == "save":
            try:
                password_to_save = value['listofpasswords'][0]
                Password_account = value['Account']
                data_to_save = [password_to_save,Password_account]
                location_to_save = value['folder_to_save']
                filename = 'mypickle.pk'
                if Password_account == "":
                    sg.popup("Please Enter Account for which you want to save the password")
                    continue
                if location_to_save == "":
                    with open(filename,'rb') as fi:
                        location_to_save = pickle.load(fi)
                        print(location_to_save)
                else : # Saving location
                    with open(filename,'wb') as fi:
                        pickle.dump(location_to_save, fi)
                        print(location_to_save)
                # Creating and Saving in Excel
                CreateExcel_try(data_to_save,location_to_save)
                window['folder_to_save'].update(value=location_to_save)
                message = f"Success! Your Password Information Stored in mentioned Location"
                window['Success'].update(value=message,text_color="Yellow")
                window['Account'].update(value="")
                continue
            except FileNotFoundError:
                sg.popup("You Must Choose the Folder First Time")
            except IndexError:
                sg.popup("Select a password to save.")

if __name__ == '__main__':
    main()
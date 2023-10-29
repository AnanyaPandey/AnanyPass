import os
import openpyxl
from pathlib import Path
import pandas as pd

path = Path('.')

def CreateExcel(rowdata,location):
    # ... (same as in your original code)
    # Check if the Excel file exists
    print("TETTING")
    target_dir = Path(location)
    os.chdir(target_dir)

    print(Path.cwd())
    print("Testing")
    file_path = 'PasswordRecords.xlsx' # Use the same file_path variable
    if os.path.exists(file_path):
        # File exists, open it
        workbook = openpyxl.load_workbook(file_path)
        # print(f"Opening existing Excel file: {file_path}")
    else:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    # Create a new worksheet (if needed)
    worksheet = workbook.active

    # Check if the first row is already filled
    first_row_empty = all(cell.value is None for cell in worksheet[1])

    if first_row_empty:
        # Add your header data to the first row
        header = ["Password","Application_Account"]
        for i, value in enumerate(header, start=1):
            worksheet.cell(row=1, column=i, value=value)
        worksheet.append(rowdata)
    else:
        # Add your data to the first row
        worksheet.append(rowdata)

    workbook.save("PasswordRecords.xlsx")
    workbook.close()
    return

def CreateExcel_try(rowdata, location):
    # Create a Path object for the target directory
    target_dir = Path(location)
    filename = 'PasswordRecords.xlsx'
    file_path = target_dir / filename
    if file_path.exists():
        # File exists, open it
        workbook = openpyxl.load_workbook(file_path)
    else:
        # Create a new Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    # Create a new worksheet (if needed)
    worksheet = workbook.active

    # Check if the first row is already filled
    first_row_empty = all(cell.value is None for cell in worksheet[1])

    if first_row_empty:
        # Add your header data to the first row
        header = ["Password", "Application_Account"]
        for i, value in enumerate(header, start=1):
            worksheet.cell(row=1, column=i, value=value)

    # Add your data to the worksheet
    worksheet.append(rowdata)

    # Save and close the workbook with the file_path
    workbook.save(file_path)
    workbook.close()

def read_from_file(xl_path = path,xl_file=Path('PasswordRecords.xlsx')) :
    completefile = xl_path/xl_file
    if completefile.exists() :
        df = pd.read_excel(xl_file)
        Listof_passwords = list(df['Password'])
        Listof_accounts = list(df['Application_Account'])
        return [Listof_passwords,Listof_accounts]
    else:
        return "Not_Found"


if __name__ == '__main__':
    R1 = ['Anany$132', 'Whatsapp']
    R2 = ['Passw0rd@@', 'Jamadar']
    R3 = ['@#2DFDf', 'Whatever']
    path = Path('C:/py/password_win_app/AnanyPass')

    read_from_file(path,'PasswordRecords.xlsx')
    #CreateExcel(R1,location)
    #CreateExcel(R2,location)
    #CreateExcel(R3,location)